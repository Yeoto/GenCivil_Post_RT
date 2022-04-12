import smtplib
from os.path import basename, splitext, isdir, isfile
from os import remove
from email.message import EmailMessage
import zipfile
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.cell import WriteOnlyCell
import pygetwindow as pw

class MyDLGLib:
    def FindAndCloseDialog(dialog_name):
        win_list = pw.getWindowsWithTitle(dialog_name)
        for win in win_list:
            win.close()
        return

class MyZiplib:
    def MakeZip(export_path:str, zip_file_list: list[str]) -> None:
        if isfile(export_path):
            remove(export_path)
        
        if len(zip_file_list) == 0:
            return

        with zipfile.ZipFile(export_path, 'w', zipfile.ZIP_DEFLATED) as file_zip:
            for file_path in zip_file_list:
                if isfile(file_path) == False:
                    continue
                file_zip.write(file_path, arcname=basename(file_path))
            file_zip.close()

class MyEmaillib:
    def Send_Report(Mail_to:list[str], message:str, attachment:list[str]=[], recursive=False) -> None:
        msg = EmailMessage()
        msg['Subject'] = 'Civil NS Post Table Regression Test Result !!'
        msg['From'] = 'pyj0827@midasit.com'

        mail_to_str = ''
        for id in Mail_to:
            if id.find('@') == -1:
                id += "@midasit.com"
            mail_to_str += id + ', '
        msg['To'] = mail_to_str[0:-2]

        msg.add_header('Content-Type', 'text')
        msg.set_payload(message)

        for file in attachment:
            if isfile(file) == False:
                continue

            with open(file, 'rb') as f:
                msg.add_attachment(f.read(), maintype='application', subtype=splitext(file)[1],filename=basename(file))

        try:
            smtp = smtplib.SMTP_SSL('smtp.mailplug.co.kr', timeout=100)
            smtp.login('pyj0827@midasit.com', 'qkrdbwls00@@')
            smtp.send_message(msg)
            smtp.quit()
        except Exception as e:
            if recursive == False:
                MyEmaillib.Send_Report(Mail_to, message + 'attachment too large. attatachment excluded', [], True)
            pass

class MyXLlib:
    __workbook = None
    __worksheet = None
    __TableName = str
    __Line = 0
    __cells = []

    def __init__(self) -> None:
        self.__workbook = Workbook(write_only=True)
        self.__TableName = ""
        self.__Line = 0

        self.__cells = []
        return

    def CreateSheet(self, TableName: str) -> None:
        if self.__worksheet != None and self.__Line == 0:
            self.__workbook.remove(self.__worksheet)

        self.__worksheet = self.__workbook.create_sheet(title=TableName)
        self.__TableName = TableName
        self.__Line = 0
        self.__cells.clear()
        return
    
    def WriteOneLine(self, datas: list[str], additional: list[str] = []) -> None:
        if len(additional) > 0:
            if len(datas) != len(additional):
                raise

        if len(self.__cells) < len(datas):
            extnedsize = len(datas) - len(self.__cells)
            for i in range(extnedsize):
                self.__cells.append(WriteOnlyCell(self.__worksheet))

        for i in range(len(datas)):
            data = datas[i]
            additional_data = ''
            if len(additional) > 0:
                additional_data = additional[i]

            cell = self.__cells[i]
            cell.value = data
            if additional_data == 'Failure':
                cell.fill = PatternFill(start_color="FFC7CE", fill_type='solid')
            elif additional_data == 'Error':
                cell.fill = PatternFill(start_color="FFEB9C", fill_type='solid')                
            else:
                cell.fill = PatternFill()

        self.__worksheet.append(self.__cells[0:len(datas)])
        self.__Line += 1
        return 

    def WriteDualLine(self, datas: list[(str, str, str)] = []) -> None:
        if len(datas) <= 0:
            return 

        if self.__Line + 2 > 1000000:
            self.WriteOneLine(['Sheet Rows are too long. See Next Sheet, Please'])
            self.CreateSheet(self.__TableName)
        
        for i in range(2):
            self.WriteOneLine(['FES' if i == 0 else 'MEC'] + [data[i].strip() for data in datas], [None] + [data[2] for data in datas])

        return 

    def WriteLine(self, datas: list[str] = [], col_offset:int = 1) -> None:
        if len(datas) <= 0:
            self.__worksheet.append([WriteOnlyCell(self.__worksheet)])
            self.__Line += 1
            return 

        if self.__Line + 1 > 1000000:
            self.WriteOneLine(['Sheet Rows are too long. See Next Sheet, Please'])
            self.CreateSheet(self.__TableName)

        self.WriteOneLine( ([None]*col_offset) + datas )
        return

    def save(self, path: str) -> None:
        if isfile(path):
            remove(path)

        if self.__worksheet != None and self.__Line == 0:
            self.__workbook.remove(self.__worksheet)

        if 'Sheet' in self.__workbook:
            self.__workbook.remove(self.__workbook['Sheet'])

        if len(self.__workbook.sheetnames) > 0:
            self.__workbook.save(path)

        self.__workbook.close()
        return 

if __name__ == "__main__":
    MyDLGLib.FindAndCloseDialog('제목 없음 - Windows 메모장')